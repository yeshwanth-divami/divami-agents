#!/usr/bin/env python3
"""
Generate scope.xlsx from a scope.tree file, using Sample-Scope.xlsx as the template.

Usage:
    python generate-scope-xlsx.py <scope.tree> <template.xlsx> <output.xlsx>

The template is copied as-is (preserving Guidelines, named ranges, and all formula
infrastructure). Only the Scope sheet's data rows are replaced.

Requirements:
    uv pip install openpyxl
"""

import re
import shutil
import sys
from pathlib import Path

# ── Abbreviation maps ──────────────────────────────────────────────────────────

COMPLEXITY = {"XS", "S", "M", "L", "XL"}

COMPONENT_MAP = {
    "FE": "Frontend",
    "BE": "Backend",
    "FS": "Fullstack",   # will write two rows: Frontend + Backend
}

FORM_FACTOR_MAP = {
    "R":  "Responsive",
    "W":  "Web",
    "MN": "Mobile Native",
    "MH": "Mobile Hybrid",
}

PHASE_MAP = {
    "P1": "Phase 1",
    "P2": "Phase 2",
    "P3": "Phase 3",
}


# ── Parser ─────────────────────────────────────────────────────────────────────

def parse_tree(path: Path) -> list[dict]:
    """
    Parse a scope.tree file into a flat list of row dicts:
      {module, feature, sub_feature, complexity, component, form_factor, phase, pages}

    Hierarchy is determined by indentation:
      [Module]           → 0 indent (square brackets)
      4-space indent     → Feature
      8-space indent     → Sub Feature  ATTRS...
    """
    rows = []
    current_module = ""
    current_feature = ""

    with open(path) as f:
        for lineno, raw in enumerate(f, 1):
            line = raw.rstrip()

            # Skip blank and comment lines
            if not line or line.lstrip().startswith("#"):
                continue

            stripped = line.lstrip()
            indent = len(line) - len(stripped)

            if stripped.startswith("[") and stripped.endswith("]"):
                # Module line
                current_module = stripped[1:-1].strip()
                current_feature = ""
                continue

            if indent == 4:
                # Feature line — no attributes
                current_feature = stripped
                continue

            if indent == 8:
                # Sub Feature line — parse attributes
                parts = stripped.split()
                if len(parts) < 5:
                    print(f"  WARNING line {lineno}: expected at least 5 tokens, got: {stripped!r}", file=sys.stderr)
                    continue

                # Name = everything before the last 4 attribute tokens
                # Attrs are always: Complexity Component FormFactor Phase [pages?]
                # Look for the complexity token to find where attrs start
                attr_start = None
                for i, p in enumerate(parts):
                    if p.upper() in COMPLEXITY:
                        attr_start = i
                        break

                if attr_start is None:
                    print(f"  WARNING line {lineno}: no complexity token found: {stripped!r}", file=sys.stderr)
                    continue

                name = " ".join(parts[:attr_start])
                attrs = parts[attr_start:]

                if len(attrs) < 4:
                    print(f"  WARNING line {lineno}: too few attributes after name: {attrs}", file=sys.stderr)
                    continue

                complexity  = attrs[0].upper()
                component   = attrs[1].upper()
                form_factor = attrs[2].upper()
                phase       = attrs[3].upper()
                pages       = int(attrs[4]) if len(attrs) > 4 else 3

                errors = []
                if complexity not in COMPLEXITY:
                    errors.append(f"unknown complexity {complexity!r}")
                if component not in COMPONENT_MAP:
                    errors.append(f"unknown component {component!r}")
                if form_factor not in FORM_FACTOR_MAP:
                    errors.append(f"unknown form_factor {form_factor!r}")
                if phase not in PHASE_MAP:
                    errors.append(f"unknown phase {phase!r}")
                if errors:
                    print(f"  ERROR line {lineno}: {', '.join(errors)}", file=sys.stderr)
                    sys.exit(1)

                base = dict(
                    module=current_module,
                    feature=current_feature,
                    sub_feature=name,
                    complexity=complexity,
                    form_factor=FORM_FACTOR_MAP[form_factor],
                    phase=PHASE_MAP[phase],
                    pages=pages,
                )

                if component == "FS":
                    rows.append({**base, "component": "Frontend"})
                    rows.append({**base, "component": "Backend"})
                else:
                    rows.append({**base, "component": COMPONENT_MAP[component]})

    return rows


# ── Formula builder ────────────────────────────────────────────────────────────

def formulas(r: int) -> dict:
    """
    Return all formula-column values for data row r (1-indexed, row 2 = first data row).
    Column letters match Sample-Scope.xlsx exactly.
    """
    return {
        # Design hours — Frontend only, VLOOKUP by form factor col 2
        "J": (
            f'=(IF(AND($G{r}="Frontend"),'
            f'VLOOKUP($F{r},Effort_Estimate_Guidelines,'
            f'VLOOKUP($H{r},FormFactorGuidelines,2,FALSE),FALSE),""))'
            f'-$X{r}'
        ),
        # Web UI — Frontend + Responsive/Web, VLOOKUP by form factor col 3
        "K": (
            f'=(IF(AND($G{r}="Frontend",'
            f'OR($H{r}="Responsive",$H{r}="Web")),'
            f'VLOOKUP($F{r},Effort_Estimate_Guidelines,'
            f'VLOOKUP($H{r},FormFactorGuidelines,3,FALSE),FALSE),""))'
            f'-$Y{r}'
        ),
        # Services — Backend only
        "L": (
            f'=(IF(OR($G{r}="Backend"),'
            f'VLOOKUP($F{r},Effort_Estimate_Guidelines,2,FALSE),""))'
            f'-$Z{r}'
        ),
        # Mobile UI — Frontend + mobile form factors
        "N": (
            f'=(IF(AND($G{r}="Frontend",'
            f'OR($H{r}="Mobile Native",$H{r}="Mobile Hybrid",$H{r}="Responsive")),'
            f'VLOOKUP($F{r},Effort_Estimate_Guidelines,'
            f'VLOOKUP($H{r},FormFactorGuidelines,3,FALSE),FALSE),""))'
            f'-$AA{r}'
        ),
        # QA Manual = 20% of (Web UI + Services)
        "T": f"=(SUM(K{r}:L{r})*0.2)",
        # QA Auto — Frontend + Responsive/Web, VLOOKUP by form factor col 4
        "U": (
            f'=(IF(AND($G{r}="Frontend",'
            f'OR($H{r}="Responsive",$H{r}="Web")),'
            f'VLOOKUP($F{r},Effort_Estimate_Guidelines,'
            f'VLOOKUP($H{r},FormFactorGuidelines,4,FALSE),FALSE),""))'
            f'-$AB{r}'
        ),
        # Validation total
        "V": f"=SUM(J{r}:U{r})",
    }


# ── Writer ─────────────────────────────────────────────────────────────────────

def write_xlsx(rows: list[dict], template: Path, output: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    shutil.copy2(template, output)
    wb = load_workbook(output)
    ws = wb["Scope"]

    # Clear existing data rows (keep header row 1)
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)

    # Style helpers
    def cell_font(col: str, row_idx: int, component: str):
        """Black for formulas, blue for inputs."""
        if col in ("B", "C", "D", "E", "F", "G", "H", "I"):
            return Font(color="0000FF")   # blue = user input
        return Font(color="000000")       # black = formula

    for i, row in enumerate(rows):
        r = i + 2   # Excel row number (1 = header)

        static = {
            "B": row["module"],
            "C": row["feature"],
            "D": row["sub_feature"],
            "E": row["pages"],
            "F": row["complexity"],
            "G": row["component"],
            "H": row["form_factor"],
            "I": row["phase"],
        }

        for col, val in static.items():
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(color="0000FF")   # blue = input

        for col, formula in formulas(r).items():
            c = ws[f"{col}{r}"]
            c.value = formula
            c.font = Font(color="000000")   # black = formula

    wb.save(output)
    print(f"Written: {output}  ({len(rows)} rows)")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) != 4:
        print("Usage: generate-scope-xlsx.py <scope.tree> <template.xlsx> <output.xlsx>")
        sys.exit(1)

    tree_path     = Path(sys.argv[1])
    template_path = Path(sys.argv[2])
    output_path   = Path(sys.argv[3])

    for p in (tree_path, template_path):
        if not p.exists():
            print(f"ERROR: not found: {p}", file=sys.stderr)
            sys.exit(1)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Parsing {tree_path} ...")
    rows = parse_tree(tree_path)
    print(f"  {len(rows)} rows parsed")

    if not rows:
        print("ERROR: no rows found in scope.tree", file=sys.stderr)
        sys.exit(1)

    print(f"Writing {output_path} ...")
    write_xlsx(rows, template_path, output_path)

    # Run recalc if available
    recalc = Path(__file__).parent.parent.parent / "excel-wrangling" / "scripts" / "recalc.py"
    if recalc.exists():
        import subprocess
        result = subprocess.run(
            [sys.executable, str(recalc), str(output_path)],
            capture_output=True, text=True,
        )
        print(result.stdout)
        if result.returncode == 2:
            print("WARNING: formula errors found — check output above", file=sys.stderr)
    else:
        print("(recalc.py not found — skipping formula validation)")


if __name__ == "__main__":
    main()
