#!/usr/bin/env python3
"""
scope-tree.py — lossless round-trip between any Excel sheet and a .tree file.

Usage:
    python scope-tree.py excel2tree <file.xlsx> <sheet> "<col1,col2,...>" [output.tree]
    python scope-tree.py tree2excel <file.tree> <template.xlsx> <output.xlsx>

    <col1,col2,...>: hierarchy columns in order (deepest = leaf node).
                     All remaining columns become tab-separated data on the leaf row.

Examples:
    python scope-tree.py excel2tree "Scope.xlsx" "Scope" "Segment,Module,Feature,Sub Feature" scope-v3.tree
    python scope-tree.py tree2excel scope-v3.tree "Scope.xlsx" scope-v3-out.xlsx

Tree format:
    # sheet: Scope
    # hierarchy: Segment | Module | Feature | Sub Feature
    # columns: Notes | #pages | Complexity | ...

    Sage
        Unified Data Layer
            Data Corpus Setup
                Static file connectors (JSON CSV Excel PDF Word)<TAB><NaN><TAB>3<TAB>M<TAB>...

Notes:
    - Column names are pipe-separated in metadata (handles commas in names).
    - Newlines inside column names are encoded as \\n in metadata.
    - None / NaN cells are stored as the literal token <NaN>.
    - tree2excel writes raw values back; formula columns are preserved from the
      template if they are NOT listed in # columns (i.e. not exported in the tree).

Requirements:
    uv pip install openpyxl
"""

import sys
from pathlib import Path
from collections import OrderedDict

NAN = "<NaN>"
INDENT = "    "  # 4 spaces per level


# ── encoding helpers ───────────────────────────────────────────────────────────

def encode_colname(name: str) -> str:
    """Encode a column name for the metadata line (escape newlines)."""
    return name.replace("\n", "\\n").replace("\r", "")


def decode_colname(name: str) -> str:
    return name.replace("\\n", "\n")


def encode_val(v) -> str:
    """Encode a cell value for storage in the tree (newlines → \\n)."""
    if v is None:
        return NAN
    s = str(v).strip().replace("\r\n", "\\n").replace("\r", "\\n").replace("\n", "\\n")
    return s if s else NAN


def decode_val(s: str):
    """Decode a stored string back to a Python value (\\n → newline)."""
    if s == NAN:
        return None
    s = s.replace("\\n", "\n")
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


# ── excel → tree ───────────────────────────────────────────────────────────────

def excel_to_tree(xlsx_path: Path, sheet_name: str, hierarchy_cols: list[str],
                  output_path: Path | None) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        _die(f"sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Read and clean header row; skip columns with no name
    raw_header = [c.value for c in ws[1]]
    header = [str(h).strip() if h is not None else "" for h in raw_header]

    # Validate hierarchy columns exist
    for col in hierarchy_cols:
        if col not in header:
            _die(f"hierarchy column {col!r} not found in header.\nHeader: {header}")

    hier_indices = [header.index(col) for col in hierarchy_cols]
    # Only include columns that have a non-empty name and are not hierarchy columns
    data_cols = [h for h in header if h and h not in hierarchy_cols]
    # Deduplicate while preserving order (in case two cols strip to same name)
    seen = set()
    deduped = []
    for h in data_cols:
        if h not in seen:
            seen.add(h)
            deduped.append(h)
    data_cols = deduped
    data_indices = [header.index(col) for col in data_cols]

    # Read all non-empty data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        hier_vals = [encode_val(row[i]) for i in hier_indices]
        data_vals = [encode_val(row[i]) for i in data_indices]
        rows.append((hier_vals, data_vals))

    depth = len(hierarchy_cols)

    # Build metadata header
    lines = [
        f"# sheet: {sheet_name}",
        f"# hierarchy: {' | '.join(encode_colname(c) for c in hierarchy_cols)}",
        f"# columns: {' | '.join(encode_colname(c) for c in data_cols)}",
        "",
    ]

    # Build tree body — emit hierarchy nodes only when they change
    prev_hier = [None] * (depth - 1)

    for hier_vals, data_vals in rows:
        for level in range(depth - 1):
            if hier_vals[level] != prev_hier[level]:
                # Reset deeper cached levels
                for deeper in range(level, depth - 1):
                    prev_hier[deeper] = None
                prev_hier[level] = hier_vals[level]
                lines.append(f"{INDENT * level}{hier_vals[level]}")

        leaf = hier_vals[-1]
        attrs = "\t".join(data_vals)
        lines.append(f"{INDENT * (depth - 1)}{leaf}\t{attrs}")

    out = "\n".join(lines) + "\n"

    if output_path:
        output_path.write_text(out, encoding="utf-8")
        print(f"Written: {output_path}  ({len(rows)} rows)")

    return out


# ── tree parser ────────────────────────────────────────────────────────────────

def parse_tree(path: Path):
    """
    Returns (sheet_name, hierarchy_cols, data_cols, rows).
    rows: list of dicts {col_name: value, ...} with decoded Python values.
    """
    sheet_name = None
    hierarchy_cols = []
    data_cols = []

    raw_lines = path.read_text(encoding="utf-8").splitlines()

    for line in raw_lines:
        if line.startswith("# sheet:"):
            sheet_name = line[len("# sheet:"):].strip()
        elif line.startswith("# hierarchy:"):
            hierarchy_cols = [decode_colname(c.strip())
                              for c in line[len("# hierarchy:"):].split("|")]
        elif line.startswith("# columns:"):
            data_cols = [decode_colname(c.strip())
                         for c in line[len("# columns:"):].split("|")]

    if not sheet_name:
        _die("tree file missing '# sheet:' metadata")
    if not hierarchy_cols:
        _die("tree file missing '# hierarchy:' metadata")

    depth = len(hierarchy_cols)
    current_hier = [None] * (depth - 1)
    rows = []

    for line in raw_lines:
        raw = line.rstrip("\n")
        stripped = raw.lstrip(" ")

        if not stripped or stripped.startswith("#"):
            continue

        indent = len(raw) - len(stripped)
        level = indent // 4

        if level < depth - 1:
            # Hierarchy (non-leaf) node
            current_hier[level] = stripped
            for deeper in range(level + 1, depth - 1):
                current_hier[deeper] = None
        else:
            # Leaf node: name TAB attr1 TAB attr2 ...
            parts = stripped.split("\t")
            leaf_name = parts[0]
            attr_vals = parts[1:]

            row = {}
            for i, col in enumerate(hierarchy_cols[:-1]):
                row[col] = current_hier[i]
            row[hierarchy_cols[-1]] = leaf_name
            for i, col in enumerate(data_cols):
                row[col] = decode_val(attr_vals[i]) if i < len(attr_vals) else None

            rows.append(row)

    return sheet_name, hierarchy_cols, data_cols, rows


# ── tree → excel ───────────────────────────────────────────────────────────────

def tree_to_excel(tree_path: Path, template_path: Path, output_path: Path) -> None:
    import shutil
    from openpyxl import load_workbook

    sheet_name, hierarchy_cols, data_cols, rows = parse_tree(tree_path)

    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    if sheet_name not in wb.sheetnames:
        _die(f"sheet {sheet_name!r} not found in template. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Read template header — build name→column-index map
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col_index = {name: idx + 1 for idx, name in enumerate(header) if name}

    missing = [c for c in hierarchy_cols + data_cols if c not in col_index]
    if missing:
        print(f"WARNING: columns not found in template header (will be skipped): {missing}",
              file=sys.stderr)

    # Clear existing data rows (keep header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    all_cols = hierarchy_cols + data_cols

    for row_idx, row_data in enumerate(rows, start=2):
        for col_name in all_cols:
            if col_name not in col_index:
                continue
            ws.cell(row=row_idx, column=col_index[col_name], value=row_data.get(col_name))

    wb.save(output_path)
    print(f"Written: {output_path}  ({len(rows)} rows)")


# ── utils ──────────────────────────────────────────────────────────────────────

def _die(msg: str):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


# ── main ───────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    if not args or args[0] not in ("excel2tree", "tree2excel"):
        print(__doc__)
        sys.exit(1)

    cmd = args[0]

    if cmd == "excel2tree":
        if len(args) < 4:
            _die("Usage: scope-tree.py excel2tree <file.xlsx> <sheet> \"<col1,col2,...>\" [output.tree]")
        xlsx_path = Path(args[1])
        sheet_name = args[2]
        hierarchy_cols = [c.strip() for c in args[3].split(",")]
        output_path = Path(args[4]) if len(args) > 4 else None

        if not xlsx_path.exists():
            _die(f"{xlsx_path} not found")

        result = excel_to_tree(xlsx_path, sheet_name, hierarchy_cols, output_path)
        if not output_path:
            print(result)

    elif cmd == "tree2excel":
        if len(args) < 4:
            _die("Usage: scope-tree.py tree2excel <file.tree> <template.xlsx> <output.xlsx>")
        tree_path = Path(args[1])
        template_path = Path(args[2])
        output_path = Path(args[3])

        for p in (tree_path, template_path):
            if not p.exists():
                _die(f"{p} not found")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        tree_to_excel(tree_path, template_path, output_path)


if __name__ == "__main__":
    main()
